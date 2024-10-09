import os;
import re;
from openpyxl import Workbook
from typing import List


strikethrough_key = "Strikethroughs"
dashsed_lines = "Dashed lines"
quoted_lists_key = "lists embedded in quotes"
open_braces_key = "Open parenthesis" 
close_braces_key = "Closing parenthesis"
figure_text_key = "Figure text"
rule_title_key = "Relevant rule"
unaffected_rules = 0

rules_problem_summary: List[dict] = []

def main():
    global unaffected_rules
    for root, dirs, files in os.walk(".\\rules"):
        for file in files:
            file_path = os.path.join(root,file)

            if file.endswith(".md"):
                with open(file_path,encoding='UTF-8') as f:
                    rule_contents = f.read()

                    rule_title = extract_title(rule_contents)    

                    rule_contents = remove_md_blocks(rule_contents)

                    rule_contents = strip_frontmatter(rule_contents)
                    rule_contents = strip_endintro(rule_contents)
                    
                    invalid_char_summary :dict = summarize_invalid_chars(rule_contents, rule_title)
                    if all(isinstance(value, str) or value == 0 for value in invalid_char_summary.values()):
                        unaffected_rules += 1
                        continue
                    print("=====================================")
                    print("file path:", file_path)
                    print(rule_title)

                    
                    rules_problem_summary.append(invalid_char_summary)
                    for key, val in invalid_char_summary.items():
                        print(f"{key}: {val}")
    issues_by_rule(rules_problem_summary)
    rules_affected_by_issue(rules_problem_summary)
    percentage_rules_affected(unaffected_rules, rules_problem_summary.__len__())
                                
def extract_title(rule_contents:str) -> str:
    title_match = re.search(r'title: (.*)', rule_contents)
    return title_match.group(1) if title_match else "no title found"

def strip_frontmatter(rule_contents:str) -> str:
    return re.sub(r'---[\s\S]*?---', "",rule_contents , count=1, flags=re.DOTALL)
    
def strip_endintro(rule_contents:str) -> str:
    return rule_contents.replace('<!--endintro-->',"")

def remove_md_blocks(rule_contents:str) -> str:
    rule_contents = re.sub(r'```mardkown[\s\S]*?```', "",rule_contents , flags=re.DOTALL)
    rule_contents = re.sub(r'```[\s\S]*?```', "",rule_contents, flags=re.DOTALL)
    return rule_contents


def summarize_invalid_chars(rule_contents:str, rule_title: str) -> dict:

    invalid_char_summary: dict = {
    }
    invalid_char_summary[rule_title_key] = rule_title
    invalid_char_summary[strikethrough_key] = len(re.findall(r'~~.*~~', rule_contents))


    dashed_lines = re.findall(r'-{2,}[^-]', rule_contents)

    
    invalid_char_summary[dashsed_lines] = len(dashed_lines)
    invalid_char_summary[quoted_lists_key] = len(re.findall(r'> \*', rule_contents)) + len(re.findall(r'>  -', rule_contents))
    invalid_char_summary[open_braces_key] = len(re.findall(r'\{', rule_contents))
    invalid_char_summary[close_braces_key] = len(re.findall(r'\}', rule_contents))
    invalid_char_summary[figure_text_key] = rule_contents.count("![Figure")

    return invalid_char_summary


def percentage_rules_affected(unaffected_rules: int, total_rules: int):
    workbook = Workbook()
    workbook_writer = workbook.active
    workbook_writer.append(["Unaffected rules", "Total rules",])
    workbook_writer.append([unaffected_rules, total_rules])
    workbook.save("percentage_rules_affected.xlsx")

def rules_affected_by_issue(rules_problem_summary: List[dict]):
    rules_affected_by_issue = {}
    issues = list(rules_problem_summary[0].keys())
    issues.remove(rule_title_key)
    for summary in rules_problem_summary:
        for issue in issues:
            if summary[issue] > 0:
                if(rules_affected_by_issue.get(issue) == None):
                    rules_affected_by_issue[issue] = 1
                    continue
                rules_affected_by_issue[issue] += 1
    workbook = Workbook()
    workbook_writer = workbook.active
    workbook_writer.append(list(rules_affected_by_issue.keys()))
    issue_occurances = []
    for key, val in rules_affected_by_issue.items():
        issue_occurances.append(val)
    workbook_writer.append(issue_occurances)
    workbook.save("rules_affected_by_issue.xlsx")
        
    
    




def issues_by_rule(rules_problem_summary: List[dict]):
    rules_problem_summary = sorted(rules_problem_summary, key=lambda x: x[rule_title_key])
    workbook = Workbook()
    workbook_writer = workbook.active
    subjects: List[str] = list(rules_problem_summary[0].keys())
    workbook_writer.append(subjects)

    for dict in rules_problem_summary:
        row_data = []
        for subject in subjects:
            row_data.append(dict[subject])
        workbook_writer.append(row_data)
    workbook.save("markdown_issues_by_rule.xlsx")



main()